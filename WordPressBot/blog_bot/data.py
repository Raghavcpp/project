import openpyxl, re, time
import indexIN
import indexAE
import indexUS

def execute_file(file_name):
    module = __import__(file_name)
    print(f"This is {file_name}.py being executed.")    
    time.sleep(3)

wb = openpyxl.load_workbook('datafile.xlsx')
ws = wb.active

website = []
title = []
content = []
categories = []
keywords = []
focus = []


for row in range(2, ws.max_row + 1):
    data = [ws.cell(row=row, column=i).value for i in range(1, ws.max_column + 1)]
    website.append(data[0])
    title.append(data[1])
    content.append(data[2])
    scategories = [part.strip() for part in data[3].split(',')]
    categories.append(scategories)
    skeywords = [part.strip() for part in data[4].split(',')]
    keywords.append(skeywords)
    focus.append(data[5])
    # print(row-2)

def india():
    print("INDIA Website")
    indexIN.indiaAjath(i,title,content,categories,keywords,focus)
    
def arab():
    print("USA Website")
    indexUS.usaAjath(i,title,content,categories,keywords,focus)

def usa():
    print("ARAB Website")
    indexAE.arabAjath(i,title,content,categories,keywords,focus)
for i in range(row-1):
    if (website[i]=="ajath.com"):
        india()
    if (website[i]=="usa.ajath.com"):
        arab()
    if (website[i]=="ajath.ae"):
        usa()
