import openpyxl, re, time
import indexIN
import indexAE
import indexUS
import gdown

# URL of the Google Sheets document
file_url = 'https://docs.google.com/spreadsheets/d/1VI5ABGgKSmFFXtNi9fycd1f9e5xquV7r/export?format=xlsx'

# Specify the destination path to save the file
destination_path = 'datafile1.xlsx'

# Download the file using gdown
gdown.download(file_url, destination_path, quiet=False)
wb = openpyxl.load_workbook(destination_path)
# wb = openpyxl.load_workbook('datafile.xlsx')
ws = wb.active

website = []
title = []
content = []
categories = []
keywords = []
focus = []


for row in range(2, ws.max_row + 1):
    data = [ws.cell(row=row, column=i).value for i in range(1, ws.max_column + 1)]
    website.append(data[0])
    title.append(data[1])
    content.append(data[2])
    scategories = [part.strip() for part in data[3].split(',')]
    categories.append(scategories)
    skeywords = [part.strip() for part in data[4].split(',')]
    keywords.append(skeywords)
    focus.append(data[5])
    # print(row-2)

def india():
    print("INDIA Website")
    indexIN.indiaAjath(i,title,content,categories,keywords,focus)
    
def usa():
    print("USA Website")
    indexUS.usaAjath(i,title,content,categories,keywords,focus)

def arab():
    print("ARAB Website")
    indexAE.arabAjath(i,title,content,categories,keywords,focus)
for i in range(row-1):
    if (website[i]=="ajath.com"):
        india()
    if (website[i]=="usa.ajath.com"):
        usa()
    if (website[i]=="ajath.ae"):
        arab()