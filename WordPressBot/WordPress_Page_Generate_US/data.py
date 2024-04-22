import openpyxl, re, time
import indexUS
import gdown

file_url = 'https://docs.google.com/spreadsheets/d/1QHROA_DlKbMvKLBETQcRFMIvFGKb3FjUO-iDVKDwdpU/export?format=xlsx'
destination_path = 'datafile.xlsx'

gdown.download(file_url, destination_path, quiet=False)
wb = openpyxl.load_workbook(destination_path)
ws = wb.active

website = []
title = []
titleSummary = []
heading = []
headingSummary = []
subheading = []
subheadingSummary = []
categories = []
graphData = []
graphLine = []
link = []

def keyvaluepair(string):
    pairs = string.split(",")
    pairs = [pair.strip() for pair in pairs]    
    key_value_pairs = [pair.split(":") for pair in pairs]
    return key_value_pairs


for row in range(2, ws.max_row + 1):
    data = [ws.cell(row=row, column=i).value for i in range(1, ws.max_column + 1)]
    website.append(data[0])
    title.append(data[1])
    titleSummary.append(data[2])
    heading.append(data[3])
    headingSummary.append(data[4])
    if data[5] is not None:
        ssubheading = [part.strip() for part in data[5].split('][')]
    else:
        pass
    subheading.append(ssubheading)
    if data[6] is not None:
        ssubheadingSummary = [part.strip() for part in data[6].split('][' or '[]')]
    else:
        pass
    subheadingSummary.append(ssubheadingSummary)
    if data[7] is not None:
        sgraphData = keyvaluepair(str(data[7]))
    else:
        pass
    graphData.append(sgraphData)
    graphLine.append(data[8])
    if data[9] is not None:
        scategories = [part.strip() for part in data[9].split(',')]
    else:
        pass
    categories.append(scategories)
    link.append(data[10])
# print(website)
# print(title)
# print(titleSummary)
# print(heading)
# print(headingSummary)
# print(subheading)
# print(subheadingSummary)
# print(graphData)
# print(link)

for i in range(row-1):
    if (website[i]=="ajath.us"):
        # print("USA Website")
        indexUS.usaAjath(i,title,titleSummary,heading,headingSummary,subheading,subheadingSummary,categories,graphData,graphLine,link)




