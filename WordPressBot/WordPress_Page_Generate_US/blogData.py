import openpyxl, re, time
import indexblogUS
import gdown

file_url = 'https://docs.google.com/spreadsheets/d/1VI5ABGgKSmFFXtNi9fycd1f9e5xquV7r/export?format=xlsx'
destination_path = 'blogfile.xlsx'

gdown.download(file_url, destination_path, quiet=False)
wb = openpyxl.load_workbook(destination_path)
ws = wb.active

website = []
title = []
content = []
categories = []
keywords = []
focus = []


for row in range(2, ws.max_row + 1):
    data = [ws.cell(row=row, column=i).value for i in range(1, ws.max_column + 1)]
    website.append(data[0])
    title.append(data[1])
    content.append(data[2])
    if data[3] is not None:
        scategories = [part.strip() for part in data[3].split(',')]
    else:
        pass
    categories.append(scategories)
    if data[3] is not None:
        skeywords = [part.strip() for part in data[4].split(',')]
    else:
        pass
    keywords.append(skeywords)
    focus.append(data[5])
    # print(row-2)

def usa():
    print("ARAB Website")
    indexblogUS.usaAjath(i,title,content,categories,keywords,focus)
for i in range(row-1):
    if (website[i]=="ajath.us"):
        usa()